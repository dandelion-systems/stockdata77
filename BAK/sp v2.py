#!/usr/bin/env python3
import sys
from json import loads
import xml.etree.ElementTree as xmlet
from openpyxl import load_workbook
from http.client import HTTPSConnection
from optparse import OptionParser

sx_list = ["NYSE", "MOEX", "SPBX"]

def get_stock_info(ticker:str, db:dict, exchange:str = "NYSE", print_progress:bool = True):
	"""
	get_stock_info(ticker, db) appends the dictionaty db with stock 
	information for ticker. conn is the active connection to the YF stock information API.
	
	The stock information is appended to db only in case db does not yet have an entry 
	with the same ticker. Otherwise, it is neither appended nor updated, which allows 
	to skip the Yahoo Finance API call.
	"""
	
	if ticker == "" or exchange not in sx_list:
		return False

	if exchange == "SPBX": exchange = "NYSE"

	if ticker in db:
		if db[ticker][0] == exchange:
			if print_progress: print("    " + ticker + " updated but API request skipped as duplicate.")
			return True

	if exchange == "NYSE":
		conn = HTTPSConnection("query1.finance.yahoo.com")
		"""
		conn is a connection point for Yahoo Finance API. The request will look like this:
		https://query1.finance.yahoo.com/v10/finance/quoteSummary/TICKER?modules=price
		Full sample output is in yf_spy_sample.json file in this folder.

		More informantion on possible calls can be found at
		https://habr.com/ru/post/505674/

		"""
		conn.request("GET", "/v10/finance/quoteSummary/" + ticker + "?modules=price")
		res = conn.getresponse()
		json_result = res.read().decode('utf-8')

		try:
			stock_info = loads(json_result)['quoteSummary']['result'][0]['price']
		except:
			if print_progress: print("    Error: probably incorrect ticker " + ticker +".")
			db[ticker] = [exchange, "Error: probably incorrect ticker " + ticker, 0.00, 0.00]
			return False
		
		try:
			if (company := stock_info['longName']) is None: company = ticker
		except:
			company = ticker
		
		try:
			price = float(stock_info['regularMarketPrice']['raw'])
		except:
			price = 0.00

		try:
			change_percent = float(stock_info['regularMarketChangePercent']['raw'])
		except:
			change_percent = 0.00
	
	elif exchange == "MOEX":
		conn = HTTPSConnection("iss.moex.com")
		"""
		https://iss.moex.com/iss/engines/stock/markets/shares/securities/AFLT.xml
		"""
		conn.request("GET", "/iss/engines/stock/markets/shares/securities/" + ticker + ".xml")
		res = conn.getresponse()
		xml_result_str = res.read().decode('utf-8')
		xml_result = xmlet.fromstring(xml_result_str)

		for dta in xml_result.findall("data"):
			if dta.attrib["id"] == "securities":
				for entry in dta.find("rows").findall("row"):
					if entry.attrib["BOARDID"] == "TQBR":
						try: 
							company = entry.attrib["SECNAME"]
						except: 
							company = ticker
						break
			if dta.attrib["id"] == "marketdata":
				for entry in dta.find("rows").findall("row"):
					if entry.attrib["BOARDID"] == "TQBR":
						try: 
							price = float(entry.attrib["LAST"])
						except: 
							price = 0.00
						try: 
							change_percent = float(entry.attrib["LASTTOPREVPRICE"]) / 100.00
						except: 
							change_percent = 0.00
						break
			

	db[ticker] = [exchange, company, price, change_percent]

	if print_progress: print("    " + ticker + " updated.")

	return True

def main():
	opt_parser = OptionParser("Usage: %prog [options] File_with_stock_tickers.xlsx")
	opt_parser.add_option("-d", "--dry", help="Only simulate, don't update the file.", action="store_true", dest='dry')

	try:
		(options,args) = opt_parser.parse_args(sys.argv)
	except:
		return

	try:
		dry = options.dry
		workbook_name = args[1]
	except:
		print(opt_parser.get_usage())
		return

	stocks = {}
	"""
	stocks is a dictionary holding the current stock information needed to update
	the Excel workbook under workbook_name.

	The structure is as follows:
	Ticker,	[Exchanger, Company Name,	Current Price,	Change to Previous Close]
	{'AAPL', 	['NYSE',	'Apple Inc.', 	157.96, 		0.02]}

	db:dict[str:[list]], 
	"""

	workbook = load_workbook(workbook_name)

	sheets_with_tickers = {}

	try:
		ctrl_sheet = workbook["#Ctrl"]
		i = 1
		sheet_name = ctrl_sheet["A" + str(i)].value
		while sheet_name != "#end of list" and sheet_name is not None:
			range1_str = str(ctrl_sheet["B" + str(i)].value)
			range1 = range1_str.split(":") if ":" in range1_str else None

			range2_str = str(ctrl_sheet["C" + str(i)].value)
			range2 = range2_str.split(":") if ":" in range2_str else None

			range3_str = str(ctrl_sheet["D" + str(i)].value)
			range3 = range3_str.split(":") if ":" in range3_str else None

			sheets_with_tickers[sheet_name] = [range1, range2, range3]
			i += 1
			sheet_name = ctrl_sheet["A" + str(i)].value
	except:
		print("Error: the workbook must have a properly formatted #Ctrl sheet.")
		return 1

	for sheet_name in sheets_with_tickers:
		print("Updating sheet " + sheet_name + ".")

		sheet = workbook[sheet_name]

		cell_range1 = sheets_with_tickers[sheet_name][0]
		cell_range2 = sheets_with_tickers[sheet_name][1]
		cell_range3 = sheets_with_tickers[sheet_name][2]

		cell_range = []
		cell_range += sheet[cell_range1[0]:cell_range1[1]] if cell_range1 is not None else []
		cell_range += sheet[cell_range2[0]:cell_range2[1]] if cell_range2 is not None else []
		cell_range += sheet[cell_range3[0]:cell_range3[1]] if cell_range3 is not None else []

		try:
			for cell_ticker in cell_range:  # sheet[] returns a tuple even if iterated, 
				cell = cell_ticker[0]       # so we need [0] to get one cell
				if (ticker := cell.value) is not None:
					ticker = ticker.upper()
					exchange = sheet.cell(cell.row, cell.column+11).value.upper()

					get_stock_info(ticker=ticker, db=stocks, exchange=exchange, print_progress=not dry)
					
					if ticker in stocks and not dry:
						sheet.cell(cell.row, cell.column-1).value = stocks[ticker][1] # update the company name
						sheet.cell(cell.row, cell.column+6).value = stocks[ticker][2] # update the current price
						sheet.cell(cell.row, cell.column+9).value = stocks[ticker][3] # update the change
					elif ticker in stocks and dry:
						trucated_name = stocks[ticker][1]
						if len(trucated_name) > 15: trucated_name = trucated_name[:15] + "..."
						trucated_name = trucated_name.ljust(20, " ")
						print(ticker.ljust(7," ") + trucated_name + "{0:8.2f} {1:5.2f}%".format(stocks[ticker][2], 100*stocks[ticker][3]))
		except:
			print("    Error: cell A1 must store the range of cells with stock symbols, for instance C4:C27. Sheet skipped.")

	workbook.save(workbook_name)

	return 0

if __name__ == '__main__':
	main()
