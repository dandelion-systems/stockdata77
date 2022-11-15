#!/usr/bin/env python3
import sys
from json import loads
from openpyxl import load_workbook
from http.client import HTTPSConnection
from optparse import OptionParser

def get_stock_info(ticker:str, db:dict, conn:HTTPSConnection, print_progress:bool = True):
	"""
	get_stock_info(ticker, db) appends the dictionaty db with stock 
	information for ticker. conn is the active connection to the YF stock information API.
	
	The stock information is appended to db only in case db does not yet have an entry 
	with the same ticker. Otherwise, it is neither appended nor updated, which allows 
	to skip the Yahoo Finance API call.
	"""
	
	if ticker == "":
		return False

	if ticker not in db:
		conn.request("GET", "/v10/finance/quoteSummary/" + ticker + "?modules=price")
		res = conn.getresponse()
		json_result = res.read().decode('utf-8')

		try:
			stock_info = loads(json_result)['quoteSummary']['result'][0]['price']
		except:
			if print_progress: print("    Error: probably incorrect ticker " + ticker +".")
			db[ticker] = ["Error: probably incorrect ticker " + ticker, 0.00, 0.00]
			return False
		
		try:
			if (company := stock_info['longName']) is None: company = ticker
		except:
			company = ticker
		
		try:
			price = float(stock_info['regularMarketPrice']['raw'])
		except:
			price = 0.00

		try:
			change_percent = float(stock_info['regularMarketChangePercent']['raw'])
		except:
			change_percent = 0.00

		db[ticker] = [company, price, change_percent]

		if print_progress: print("    " + ticker + " updated.")
	else:
		if print_progress: print("    " + ticker + " updated but YF request skipped as duplicate.")

	return True

def main():
	opt_parser = OptionParser("Usage: %prog [options] File_with_stock_tickers.xlsx")
	opt_parser.add_option("-d", "--dry", help="Only simulate, don't update the file.", action="store_true", dest='dry')

	try:
		(options,args) = opt_parser.parse_args(sys.argv)
	except:
		return

	try:
		dry = options.dry
		workbook_name = args[1]
	except:
		print(opt_parser.get_usage())
		return

	#workbook_name = "Stocks.xlsx"
	"""
	workbook_name stores the name of a preformatted Excel workbook that
	will be updated with stock information.
	"""

	apiConn = HTTPSConnection("query1.finance.yahoo.com")
	"""
	apiConn is a connection point for Yahoo Finance API. The request will look like this:
	https://query1.finance.yahoo.com/v10/finance/quoteSummary/TICKER?modules=price
	Full sample output is in yf_spy_sample.json file in this folder.

	More informantion on possible calls can be found at
	https://habr.com/ru/post/505674/

	"""

	stocks = {}
	"""
	stocks is a dictionary holding the current stock information needed to update
	the Excel workbook under workbook_name.

	The structure is as follows:
	Ticker,	[Company Name,	Current Price,	Change to Previous Close]
	{'AAPL', 	['Apple Inc.', 	157.96, 		0.02]}

	db:dict[str:[list]], 
	"""

	workbook = load_workbook(workbook_name)

	for sheet_name in workbook.sheetnames:
		
		# Cell A1 must hold the range of cells with tickers in 'C1:C20' like format
		sheet = workbook[sheet_name]
		ticker_range = sheet['A1'].value 
		
		if ticker_range == None:
			print("Skipping sheet " + sheet_name + ".")
			continue

		print("Updating sheet " + sheet_name + ".")

		try:
			r = ticker_range.split(":")
			for tickerCell in sheet[r[0]:r[1]]:  # sheet[] returns a tuple even if iterated
				cell = tickerCell[0]			 # so, a cell is extracted from it as entry 0
				if (ticker := cell.value) is not None:
					ticker = ticker.upper()
				else:
					ticker = ""
				
				if dry: 
					print_progress = False
				else:
					print_progress = True

				get_stock_info(ticker=ticker, db=stocks, conn=apiConn, print_progress=print_progress)
				
				if ticker in stocks and not dry:
					sheet.cell(cell.row, cell.column-1).value = stocks[ticker][0] # update the company name
					sheet.cell(cell.row, cell.column+6).value = stocks[ticker][1] # update the current price
					sheet.cell(cell.row, cell.column+9).value = stocks[ticker][2] # update the change
				elif ticker in stocks and dry:
					trucated_name = stocks[ticker][0]
					if len(trucated_name) > 10: trucated_name = trucated_name[:10] + "..."
					trucated_name = trucated_name.ljust(15, " ")
					print(ticker.ljust(7," ") + trucated_name + "{0:8.2f} {1:5.2f}%".format(stocks[ticker][1], 100*stocks[ticker][2]))
		except:
			print("    Error: cell A1 must store the range of cells with stock symbols, for instance C4:C27. Sheet skipped.")

	workbook.save(workbook_name)

	return 0

if __name__ == '__main__':
	main()
