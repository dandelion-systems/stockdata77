#!/usr/bin/env python3

import sys
import os
from openpyxl import load_workbook
from optparse import OptionParser
from stockquotes import Stocks

def main():
	optParser = OptionParser(usage="Usage: %prog [options] file_with_stock_tickers.xlsx" + os.linesep + "See --help for options.")
	optParser.add_option("-d", "--dry", help="only simulate, don't update the file", action="store_true", dest='dry')

	try:
		(options,args) = optParser.parse_args(sys.argv)
		dry = options.dry
		workbookName = args[1]
	except IndexError:				# handle unknown options as OptionParser does not
		optParser.print_help()		# print help by default in this case
		return
	except:							# if --help is specified, OptionParser prints help by defaut
		return						# and throws an exception. We handle it gracefully here

	workbook = load_workbook(workbookName)

	sheetsWithTickers = {}

	try:
		ctrlSheet = workbook["#Ctrl"]
		i = 1
		sheetName = ctrlSheet["A" + str(i)].value
		while sheetName != "#end of list" and sheetName is not None:
			range1_str = str(ctrlSheet["B" + str(i)].value)
			range1 = range1_str.split(":") if ":" in range1_str else None

			range2_str = str(ctrlSheet["C" + str(i)].value)
			range2 = range2_str.split(":") if ":" in range2_str else None

			range3_str = str(ctrlSheet["D" + str(i)].value)
			range3 = range3_str.split(":") if ":" in range3_str else None

			sheetsWithTickers[sheetName] = [range1, range2, range3]

			i += 1
			sheetName = ctrlSheet["A" + str(i)].value
	except:
		print("Error: the workbook must have a properly formatted #Ctrl sheet.")
		return

	stockDB = Stocks()

	for sheetName in sheetsWithTickers:
		if not dry:
			print("Updating sheet " + sheetName + ".")

		sheet = workbook[sheetName]

		cellRange1 = sheetsWithTickers[sheetName][0]
		cellRange2 = sheetsWithTickers[sheetName][1]
		cellRange3 = sheetsWithTickers[sheetName][2]

		cellRange = []
		cellRange += sheet[cellRange1[0]:cellRange1[1]] if cellRange1 is not None else []
		cellRange += sheet[cellRange2[0]:cellRange2[1]] if cellRange2 is not None else []
		cellRange += sheet[cellRange3[0]:cellRange3[1]] if cellRange3 is not None else []

		for cellWithTicker in cellRange:   # sheet[] returns a tuple even if iterated, 
			cell = cellWithTicker[0]       # so we need [0] to get one cell
		
			ticker = cell.value
			api = sheet.cell(cell.row, cell.column+11).value

			if (key := stockDB.append(ticker, api)) is not None and not dry:
				sheet.cell(cell.row, cell.column-1).value = stockDB.getCompanyName(key) # update the company name
				sheet.cell(cell.row, cell.column+6).value = stockDB.getPrice(key) 		# update the current price
				sheet.cell(cell.row, cell.column+9).value = stockDB.getPriceChng(key) 	# update the change

	if dry:
		print(stockDB)
	else:
		workbook.save(workbookName)

if __name__ == "__main__":
	try:
		main()
	except:
		print (sys.exc_info())
